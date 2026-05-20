"""
Módulo de Importação/Exportação de Grupos
Responsável por: upload de Excel, validações, processamento, exportação de relatórios
"""

import io
import json
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .sheets import fetch_grupos, atualizar_grupo_sheets, criar_grupo


# ═════════════════════════════════════════════════════════════════════════════
# VALIDAÇÃO E PARSING
# ═════════════════════════════════════════════════════════════════════════════

COLUNAS_OBRIGATORIAS = {
    "adm": str,
    "grupo": str,
    "tipo_bem": str,
    "maior_credito": float,
    "menor_credito": float,
    "taxa_adm": float,
}

COLUNAS_OPCIONAIS = {
    "fundo_rsv": float,
    "investidor": float,
    "conservador_24m": float,
    "moderado_12m": float,
    "status": str,
    "observacoes": str,
}

TODAS_COLUNAS = {**COLUNAS_OBRIGATORIAS, **COLUNAS_OPCIONAIS}


def validar_arquivo_excel(arquivo_bytes: bytes) -> Tuple[bool, str]:
    """Validar se arquivo é Excel válido."""
    try:
        wb = load_workbook(io.BytesIO(arquivo_bytes))
        ws = wb.active
        if not ws or ws.max_row == 0:
            return False, "Arquivo vazio ou sem abas"
        return True, ""
    except Exception as e:
        return False, f"Erro ao ler arquivo: {str(e)}"


def extrair_dados_excel(arquivo_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """
    Extrair dados de arquivo Excel usando openpyxl.

    Retorna:
    - Lista de dicts com dados
    - Lista de erros encontrados
    """
    erros = []
    dados = []

    try:
        wb = load_workbook(io.BytesIO(arquivo_bytes))
        ws = wb.active

        if not ws or ws.max_row < 2:
            erros.append("Arquivo vazio ou sem linhas de dados")
            return [], erros

        # Ler header (primeira linha)
        header = []
        for cell in ws[1]:
            header.append(str(cell.value).lower() if cell.value else "")

        # Detectar colunas automaticamente (case-insensitive)
        colunas_arquivo = {col.lower(): (col, idx) for idx, col in enumerate(header)}

        # Processar linhas (começar da linha 2, pulando header)
        for linha_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                linha_data = {}

                # Processar cada coluna obrigatória
                for col_esperada, col_tipo in COLUNAS_OBRIGATORIAS.items():
                    if col_esperada not in colunas_arquivo:
                        erros.append(f"Linha {linha_num}: coluna obrigatória '{col_esperada}' não encontrada")
                        raise ValueError(f"Coluna obrigatória ausente: {col_esperada}")

                    col_nome, col_idx = colunas_arquivo[col_esperada]
                    valor = row[col_idx].value if col_idx < len(row) else None

                    # Validar tipo e converter
                    if valor is None or valor == "":
                        erros.append(f"Linha {linha_num}: '{col_esperada}' está vazio")
                        raise ValueError(f"Campo obrigatório vazio: {col_esperada}")

                    try:
                        if col_tipo == float:
                            linha_data[col_esperada] = float(valor)
                        elif col_tipo == str:
                            linha_data[col_esperada] = str(valor).strip()
                    except ValueError:
                        erros.append(f"Linha {linha_num}: '{col_esperada}' tipo inválido (esperado {col_tipo.__name__})")
                        raise ValueError(f"Tipo inválido para {col_esperada}")

                # Processar colunas opcionais
                for col_esperada, col_tipo in COLUNAS_OPCIONAIS.items():
                    if col_esperada in colunas_arquivo:
                        col_nome, col_idx = colunas_arquivo[col_esperada]
                        valor = row[col_idx].value if col_idx < len(row) else None

                        if valor is not None and valor != "":
                            try:
                                if col_tipo == float:
                                    linha_data[col_esperada] = float(valor)
                                elif col_tipo == str:
                                    linha_data[col_esperada] = str(valor).strip()
                            except ValueError:
                                erros.append(f"Linha {linha_num}: '{col_esperada}' tipo inválido")

                # Validações de negócio
                if linha_data["maior_credito"] < linha_data["menor_credito"]:
                    erros.append(f"Linha {linha_num}: maior_credito < menor_credito")
                    raise ValueError("Créditos inválidos")

                if linha_data["taxa_adm"] < 0 or linha_data["taxa_adm"] > 100:
                    erros.append(f"Linha {linha_num}: taxa_adm deve estar entre 0-100")
                    raise ValueError("Taxa inválida")

                dados.append(linha_data)

            except ValueError:
                # Pular linha com erro
                continue

        return dados, erros

    except Exception as e:
        erros.append(f"Erro geral ao processar Excel: {str(e)}")
        return [], erros


def validar_schema(dados: List[Dict]) -> Tuple[bool, List[str]]:
    """Validar schema completo dos dados."""
    erros = []

    if not dados:
        erros.append("Nenhuma linha válida para processar")
        return False, erros

    # Validações globais
    adms_unicas = set(d["adm"].upper() for d in dados)
    grupos_por_adm = {}

    for d in dados:
        adm = d["adm"].upper()
        grupo = d["grupo"]

        if adm not in grupos_por_adm:
            grupos_por_adm[adm] = set()

        if grupo in grupos_por_adm[adm]:
            erros.append(f"Duplicação detectada: ADM '{adm}' já tem grupo '{grupo}'")

        grupos_por_adm[adm].add(grupo)

    return len(erros) == 0, erros


def preview_importacao(dados: List[Dict], limite: int = 10) -> Dict:
    """
    Gerar preview dos dados a importar.

    Retorna:
    - preview: primeiras N linhas
    - total: total de linhas
    - colunas: lista de colunas detectadas
    """
    return {
        "preview": dados[:limite],
        "total": len(dados),
        "colunas": list(dados[0].keys()) if dados else [],
        "limite_preview": limite,
        "tem_mais": len(dados) > limite
    }


def processar_importacao(dados: List[Dict], modo: str = "insert_update") -> Dict:
    """
    Processar importação de dados.

    Modos:
    - "insert_update": atualizar existentes, inserir novos
    - "insert": apenas inserir, pular duplicatas
    - "update": apenas atualizar existentes

    Retorna:
    - inseridos: qtd de novos grupos
    - atualizados: qtd de grupos atualizados
    - erros: lista de erros durante processamento
    """
    grupos_existentes = fetch_grupos()
    map_existentes = {(g["adm"].upper(), g["grupo"]): g for g in grupos_existentes}

    inseridos = 0
    atualizados = 0
    erros = []

    for linha in dados:
        try:
            adm = linha["adm"].upper()
            grupo = linha["grupo"]
            chave = (adm, grupo)

            if chave in map_existentes:
                # Grupo já existe
                if modo in ["insert_update", "update"]:
                    # Atualizar
                    atualizar_grupo_sheets(
                        grupo_id=grupo,
                        adm=adm,
                        dados=linha
                    )
                    atualizados += 1
                elif modo == "insert":
                    # Pular (já existe)
                    pass
            else:
                # Grupo novo
                if modo in ["insert_update", "insert"]:
                    # Inserir
                    criar_grupo(
                        adm=adm,
                        grupo=grupo,
                        dados=linha
                    )
                    inseridos += 1

        except Exception as e:
            erros.append({
                "grupo": linha.get("grupo", "?"),
                "adm": linha.get("adm", "?"),
                "erro": str(e)
            })

    return {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "erros": erros,
        "sucesso": len(erros) == 0,
        "timestamp": datetime.now().isoformat()
    }


# ═════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO
# ═════════════════════════════════════════════════════════════════════════════

def exportar_excel_completo() -> bytes:
    """Exportar todos os grupos em Excel completo usando openpyxl."""
    grupos = fetch_grupos()

    # Definir ordem de colunas
    colunas_ordem = list(COLUNAS_OBRIGATORIAS.keys()) + list(COLUNAS_OPCIONAIS.keys())

    # Se há grupos, garantir que todas as colunas existem
    if grupos:
        todas_chaves = set()
        for g in grupos:
            todas_chaves.update(g.keys())
        colunas_presentes = [c for c in colunas_ordem if c in todas_chaves]
        colunas_presentes += [c for c in todas_chaves if c not in colunas_presentes]
    else:
        colunas_presentes = colunas_ordem

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grupos"

    # Escrever header
    for col_idx, col_nome in enumerate(colunas_presentes, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_nome)
        cell.fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Escrever dados
    for row_idx, grupo in enumerate(grupos, start=2):
        for col_idx, col_nome in enumerate(colunas_presentes, start=1):
            valor = grupo.get(col_nome, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar largura colunas
    for col_idx, col_nome in enumerate(colunas_presentes, start=1):
        max_length = len(str(col_nome))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            try:
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Converter para bytes
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io.getvalue()


def exportar_por_adm(adm: str) -> bytes:
    """Exportar grupos filtrados por administradora usando openpyxl."""
    grupos = fetch_grupos()
    grupos_filtrados = [g for g in grupos if g["adm"].upper() == adm.upper()]

    colunas_ordem = list(COLUNAS_OBRIGATORIAS.keys()) + list(COLUNAS_OPCIONAIS.keys())

    if grupos_filtrados:
        todas_chaves = set()
        for g in grupos_filtrados:
            todas_chaves.update(g.keys())
        colunas_presentes = [c for c in colunas_ordem if c in todas_chaves]
        colunas_presentes += [c for c in todas_chaves if c not in colunas_presentes]
    else:
        colunas_presentes = colunas_ordem

    # Criar workbook com nome da ADM (max 31 chars para aba)
    wb = Workbook()
    ws = wb.active
    ws.title = adm[:31]

    # Escrever header
    for col_idx, col_nome in enumerate(colunas_presentes, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_nome)
        cell.fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Escrever dados
    for row_idx, grupo in enumerate(grupos_filtrados, start=2):
        for col_idx, col_nome in enumerate(colunas_presentes, start=1):
            valor = grupo.get(col_nome, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar largura colunas
    for col_idx, col_nome in enumerate(colunas_presentes, start=1):
        max_length = len(str(col_nome))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            try:
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Converter para bytes
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io.getvalue()


def exportar_grupo(grupo_id: str, adm: str) -> Dict:
    """Exportar detalhes completos de um grupo."""
    grupos = fetch_grupos()
    grupo = next((g for g in grupos if g["grupo"] == grupo_id and g["adm"].upper() == adm.upper()), None)

    if not grupo:
        return {"erro": f"Grupo '{grupo_id}' da ADM '{adm}' não encontrado"}

    return {
        "sucesso": True,
        "grupo": grupo,
        "timestamp": datetime.now().isoformat()
    }


def exportar_relatorio_adms() -> bytes:
    """Exportar relatório comparativo de administradoras usando openpyxl."""
    grupos = fetch_grupos()

    # Agrupar por ADM
    relatorio = {}
    for g in grupos:
        adm = g["adm"]
        if adm not in relatorio:
            relatorio[adm] = {
                "total_grupos": 0,
                "credito_total": 0,
                "credito_medio": 0,
                "taxa_adm_media": 0,
                "lance_maior_medio": 0,
                "grupos": []
            }

        relatorio[adm]["total_grupos"] += 1
        relatorio[adm]["credito_total"] += g.get("maior_credito", 0)
        relatorio[adm]["taxa_adm_media"] += g.get("taxa_adm", 0)
        relatorio[adm]["grupos"].append(g["grupo"])

    # Calcular médias
    for adm in relatorio:
        total = relatorio[adm]["total_grupos"]
        if total > 0:
            relatorio[adm]["credito_medio"] = relatorio[adm]["credito_total"] / total
            relatorio[adm]["taxa_adm_media"] = relatorio[adm]["taxa_adm_media"] / total

    # Montar dados relatório
    dados_relatorio = []
    for adm, dados in relatorio.items():
        dados_relatorio.append({
            "Administradora": adm,
            "Total de Grupos": dados["total_grupos"],
            "Crédito Total (R$)": dados["credito_total"],
            "Crédito Médio (R$)": dados["credito_medio"],
            "Taxa ADM Média (%)": dados["taxa_adm_media"]
        })

    # Ordenar por Total de Grupos (descendente)
    dados_relatorio.sort(key=lambda x: x["Total de Grupos"], reverse=True)

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório ADMs"

    # Colunas do relatório
    colunas_rel = ["Administradora", "Total de Grupos", "Crédito Total (R$)", "Crédito Médio (R$)", "Taxa ADM Média (%)"]

    # Escrever header
    for col_idx, col_nome in enumerate(colunas_rel, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_nome)
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Escrever dados
    for row_idx, dados in enumerate(dados_relatorio, start=2):
        for col_idx, col_nome in enumerate(colunas_rel, start=1):
            valor = dados.get(col_nome, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar largura colunas
    for col_idx, col_nome in enumerate(colunas_rel, start=1):
        max_length = len(str(col_nome))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            try:
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Converter para bytes
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io.getvalue()
